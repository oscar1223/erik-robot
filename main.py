import requests
import csv
import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

# CONFIGURACIÓN
API_URL = "https://ejemplo.com/api/coordinates"  # ← Cambia por tu URL real
INTERVALO_SEGUNDOS = 1

# Obtener datos desde la API
def obtener_coordenada():
    try:
        response = requests.get(API_URL)
        response.raise_for_status()
        data = response.json()
        data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return data
    except Exception as e:
        print(f"Error al obtener datos: {e}")
        return None

# Guardar en CSV
def guardar_en_csv(coordenada, csv_file):
    try:
        headers = list(coordenada.keys())
        archivo_nuevo = not os.path.exists(csv_file)
        with open(csv_file, mode='a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            if archivo_nuevo:
                writer.writeheader()
            writer.writerow(coordenada)
        print(f"[CSV] Coordenada guardada")
    except Exception as e:
        print(f"Error al guardar en CSV: {e}")

# Guardar en Excel
def guardar_en_excel(coordenada, excel_file):
    try:
        headers = list(coordenada.keys())

        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Coordenadas"
            ws.append(headers)

        ws.append([coordenada.get(h, "") for h in headers])
        wb.save(excel_file)
        print(f"[Excel] Coordenada guardada")
    except Exception as e:
        print(f"Error al guardar en Excel: {e}")

# Bucle principal
def main():
    print("Recolectando coordenadas cada 1 segundo... (Ctrl+C para detener)")
    try:
        while True:
            fecha = datetime.now().strftime("%Y-%m-%d")
            csv_file = f"coordenadas_{fecha}.csv"
            excel_file = f"coordenadas_{fecha}.xlsx"

            coordenada = obtener_coordenada()
            if coordenada:
                guardar_en_csv(coordenada, csv_file)
                guardar_en_excel(coordenada, excel_file)
            else:
                print("No se recibió ninguna coordenada.")

            time.sleep(INTERVALO_SEGUNDOS)
    except KeyboardInterrupt:
        print("\nRecolección detenida por el usuario.")

if __name__ == "__main__":
    main()
